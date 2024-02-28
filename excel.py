from openpyxl import load_workbook


def get_info_from_file(path):
    wb = load_workbook(path)
    rows = []
    for sheet in wb:
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
            if not all([value == None for value in row]):
                rows.append(row)
    return rows
